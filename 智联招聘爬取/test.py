import requests
import re
import json
import random
from openpyxl import Workbook
import pymysql
import pymongo
from openpyxl.styles import Alignment


class ZhiLian:
    def __init__(self):
        self.headers = {
            'authority': 'sou.zhaopin.com',
            'cache-control': 'max-age=0',
            'upgrade-insecure-requests': '1',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.198 Safari/537.36',
            'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
            'sec-fetch-site': 'same-origin',
            'sec-fetch-mode': 'navigate',
            'sec-fetch-user': '?1',
            'sec-fetch-dest': 'document',
            'referer': 'https://www.zhaopin.com/',
            'accept-language': 'zh-CN,zh;q=0.9',
            'cookie': 'x-zp-client-id=6fda8c91-c692-4072-bb05-ed7821729d7b; sts_deviceid=1740685217a304-0c8ceeab8b02e1-376b4502-1049088-1740685217b380; FSSBBIl1UgzbN7N443S=mN0egr.AUkniHvAZ5nG8WWHQ0Bw0OVyjLVzUbidRS202qIole_qWLwRQ4f92VeI.; _uab_collina=160238182376662700063028; ssxmod_itna=YqUOAKDK7KY5D5P0dD=G7mFG=E23/jC=Qmpx0vcheGzDAxn40iDtoxTZmY=nEliBm+POFAruo48zmRYRx=AC32WTDU4i8DCkro3bDee=D5xGoDPxDeDADYo6DAqiOD7k=DEDmb8DYxGAnKqDgDYQDGuPjD7QDId==So=V0bigWeUPlu4dbqDM7eGXtDnebu0i5=hltOQDzLaDtwtgbM=Px0Pzm8GX9OG5nZmPQBDrzAhPY++zYYx4zYvbSbq3S4AKm756uFD; _jzqx=1.1606138603.1606138603.1.jzqsr=so%2Ecom|jzqct=/link.-; __xsptplus30=30.1.1606138605.1606138605.1%233%7Cwww.so.com%7C%7C%7C%7C%23%23ghY_CLKg-USQDqYLXi9hfovZT_0785Jw%23; adfbid2=0; adfbid=0; sts_sg=1; sts_sid=177d25565451df-0dedf1d13adde9-3e604809-1049088-177d2556546383; sts_chnlsid=121122523; zp_src_url=https%3A%2F%2Fwww.baidu.com%2Fbaidu.php%3Fsc.K00000K3Zd4fCW_uECt_euLayZfyEhZZ_0teD--NdHlS7wlorI_ehqQrzr9QAdzYo5Ucz4VnVF-SsbBqlEopOW27KwhP6yW6IPTyFmywMduWLny4O2PRxHe924de5GZsLZwc1Srww_QBi6rW1lTwe4G4mbXcUyi2kf-fvhL38EPqXodHxH3owhNfVvET07XPvMIszVnVbxdv-IQ9YG02tki_RNQq.7R_NR2Ar5Od669BCXgjRzeASFDZtwhUVHf632MRRt_Q_DNKnLeMX5DkgboozuPvHWdsHRy2J7jZZOlsfRymoM4EQ9JuIWxDBaurGtIKnLxKfYt_U_DY2yQvTyjtLsqT7jHzlRL5spy59OPt5gKfYtVKnv-WF_tUQQrPMgKfYt_QCJamJj7jQdsRP5Qa1Gk_EdwnmtxyrlAWv3h5USEcSEIe5-g8zxJgOWtMSLqpN2s1f_I-hzEBC.U1Yz0ZDqd_xKJ_L3dIjA8nL30ZKGm1Yk0Zfqd_xKJVUZspoNYnp3dIjA80KGUHYznWR0u1dEuZCk0ZNG5yF9pywd0ZKGujYkn6KWpyfqrHm0UgfqnH0krNtknjDLg1csPH7xn10sn-t1PW0k0AVG5H00TMfqPHfs0AFG5HDdr7tznjwxnWDLg1RsnsKVm1Yknj0kg1D3P1b4njRvPj7xnW0dnNtknjFxn0KkTA-b5H00TyPGujYs0ZFMIA7M5H00mycqn7ts0ANzu1Yz0ZKs5H00UMus5H08nj0snj0snj00Ugws5H00uAwETjYs0ZFJ5H00uANv5gKW0AuY5H00TA6qn0KET1Ys0AFL5HDs0A4Y5H00TLCq0A71gv-bm1dsTzdMXh410A-bm1dcHbc0IA7zuvNY5Hm1g1KxnHR10ZwdT1YLrHDdPWRYnW03P164nWnsPHDz0ZF-TgfqnHmkPjDYnHRYnjTsP6K1pyfqmyFWuH7WnHfsnj0sP1FbmsKWTvYqrDF7rDczfYcdfbcLfHFDPfK9m1Yk0ZK85H00TydY5H00Tyd15H00XMfqn0KVmdqhThqV5HKxn7tsg1Kxn0Kbmy4dmhNxTAk9Uh-bT1Ysg1Kxn7t1n1mYP1RYg100TA7Ygvu_myTqn0Kbmv-b5H00ugwGujYVnfK9TLKWm1Ys0ZNspy4Wm1Ys0Z7VuWYs0AuWIgfqn0KGTvP_5H00XMK_Ignqn0K9uAu_myTqnfK_uhnqn0KbmvPb5RDvPj7aPH7anjN7PHnzPjI7wj97wjRdPbmkPjcvwbDdxjfLn1DLPHR1njDsP1bvr0KYTh7buHYLPW0znjc0mhwGujdKnWPKnWD4fWnLwH0zwDfYPDDsnHKjnbfzrHm1wHDknsKEm1Yk0AFY5H00Uv7YI1Ys0AqY5HD0ULFsIjYkc10Wnznzc1cdrjbzn1cYc1cdnj0WnWRsna3snj0snj0Wninzc10WQinsQW0snj0snankQW0snj0sn0K3TLwd5HcLPHTYP1T0TNqv5H08rj-xna3sn7tsQW0sg108nW7xna3zPdtsQWc3g1Dsna3sn7ts0AF1gLKzUvwGujYs0A-1gvPsmHYs0APs5H00mLFW5HmzrHD1%26xst%3Dm1dKPWfkfWRkfW0dwHR1nWfLwRf3wRfdPHuAnHfzPbuKPgsYP1nkP1Rdn10knjT4PW6KmWdKnWPKnWD4fWnLwH0zwDfYPDDsnHKjnbfzrHm1wHDkns715HDzn1T4rjm1nWbLrHD1njcdnWfYg1czPNtk0gTqd_xKJVUZspoNYnp3dIjA807k5IUZspoPSPgfkoWPS07d5HcLPHTYP1TKIjYkPWDYnHfkPHfk0ydk5H0an0cV0yPC5yuWgLKW0Hf4P1T4Pjfkns%26word%3D%26ck%3D8049.4.103.294.400.315.412.187%26shh%3Dwww.baidu.com%26us%3D1.0.1.0.2.1231.0%26wd%3D%26bc%3D110101; urlfrom=121114583; urlfrom2=121114583; adfcid=www.baidu.com; adfcid2=www.baidu.com; Hm_lvt_38ba284938d5eddca645bb5e02a02006=1614141560; acw_tc=2760829316141415631111832e5cf21b6cdc07f017ce204ad4beddf285cf89; locationInfo_search={%22code%22:%22795%22%2C%22name%22:%22%E8%B4%BA%E5%B7%9E%22%2C%22message%22:%22%E5%8C%B9%E9%85%8D%E5%88%B0%E5%B8%82%E7%BA%A7%E7%BC%96%E7%A0%81%22}; 1420ba6bb40c9512e9642a1f8c243891=088674a2-afe5-4b78-9949-9f843e0d7eba; at=2b7d15c058c24c88a9c57897fa4b5b2f; rt=7b0a8d62cd9c4885a778fbcde4617a43; ZP_OLD_FLAG=false; sensorsdata2015jssdkcross=%7B%22distinct_id%22%3A%221087029882%22%2C%22first_id%22%3A%221775dda49e31d8-03523d0cac2fb4-3e604809-1049088-1775dda49e40%22%2C%22props%22%3A%7B%22%24latest_traffic_source_type%22%3A%22%E8%87%AA%E7%84%B6%E6%90%9C%E7%B4%A2%E6%B5%81%E9%87%8F%22%2C%22%24latest_search_keyword%22%3A%22%E6%9C%AA%E5%8F%96%E5%88%B0%E5%80%BC%22%2C%22%24latest_referrer%22%3A%22https%3A%2F%2Fwww.baidu.com%2Flink%22%2C%22%24latest_utm_source%22%3A%22baiduPC%22%2C%22%24latest_utm_medium%22%3A%22CPC%22%2C%22%24latest_utm_campaign%22%3A%22pp%22%2C%22%24latest_utm_content%22%3A%22tj%22%2C%22%24latest_utm_term%22%3A%2228700030%22%7D%2C%22%24device_id%22%3A%221775dda49e31d8-03523d0cac2fb4-3e604809-1049088-1775dda49e40%22%7D; LastCity%5Fid=530; LastCity=%E5%8C%97%E4%BA%AC; ZL_REPORT_GLOBAL={%22/resume/new%22:{%22actionid%22:%22ff4a93eb-2b8a-44f4-808d-9a7692ca3144%22%2C%22funczone%22:%22addrsm_ok_rcm%22}%2C%22//www%22:{%22seid%22:%222b7d15c058c24c88a9c57897fa4b5b2f%22%2C%22actionid%22:%22410790a0-609f-49c8-ab41-b1d3cee856cf-cityPage%22}}; sts_evtseq=12; Hm_lpvt_38ba284938d5eddca645bb5e02a02006=1614141629; FSSBBIl1UgzbN7N443T=5FkIJ1sCujhAIVtL0u59V_o8r1BqcyIy2WGLhcR0zIvsaar6vgvlx7SAwPwndo8lAwrosGW1vFrFK.oDRL3P19YipuDr9aDhwfk_5cgAGvl0XGH3063K2ylkpvjTChdOOKTCImh2EvGD82IWtqZJJVQLCQ63VvIB0bIIYf9MijEkXe4NONyNjubdy.ttgQ00RdBroaDiJNHWNwHbzeue_LcLtZ9kJZVAtZlG3dJNDqKYz7S9DNDEJKPiddr49BQAUaTZchKYUysUtbLO58PgCLuMP1IGVEgCfYCj.aHTWGhJQR.q4xqgESlTEwurl.au3onN2F49_cu4jFwPipKacMMhr',
        }
        self.user_agent = [
            "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.198 Safari/537.36",
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.88 Safari/537.36',
            "Mozilla/5.0 (SymbianOS/9.4; Series60/5.0 NokiaN97-1/20.0.019; Profile/MIDP-2.1 Configuration/CLDC-1.1) AppleWebKit/525 (KHTML, like Gecko) BrowserNG/7.1.18124",
            "Mozilla/5.0 (compatible; MSIE 9.0; Windows Phone OS 7.5; Trident/5.0; IEMobile/9.0; HTC; Titan)",
            "Mozilla/4.0 (compatible; MSIE 6.0; ) Opera/UCWEB7.0.2.37/28/999",
        ]
        self.url = 'https://sou.zhaopin.com/?jl=530&kw=Python&p={}'
        self.json_url = "https://fe-api.zhaopin.com/c/i/jobs/position-detail-new?at=4c5e59ec8ed140fcb27574fd34e97dfd" \
                        "&rt=d30c1b8f34c048edb59e5684b4ccb35e&number={" \
                        "}&_v=0.41053085&x-zp-page-request-id=51e20a9823aa4fc4bc2b1d347d7aaa59-1614259088304-624518&x" \
                        "-zp-client-id=6fda8c91-c692-4072-bb05-ed7821729d7b&MmEwMD=5to8JcU6u_HpIliN0O_2VjkIrc" \
                        ".rcNMg2Q9Nh1wTzHO1aGmCvLOmxZTpwnRudUFmAJmDsaEsvtmMK4koRgQO10pHp7CaMSz9RS.v7hoZhPbdfIvbTMG5i" \
                        ".nzBZZSYHU8FVk7VnDwMu0K8B1AJbuPRCzc5gpnIf3ZhslNGXYWae1L_cVV7LtND1mewC8ZVXD26Pxw6wA_5CXKe08yPsZve44uADE4CroKO.28_jhItv48qs8oxPouPrqS8AtdhgUnt0uhC21VvaoaVZyU_p4UmMaHj20joq9lHnTDQT0wz8VkEGsmz_AiY_nh5_1qGrwhxjivmXg0zVlK7XCrzRDHjoYglC0Y8oAk1Wjzx5ZlLnHvseZJaiej_xggNTe.pN4hudgGJKMxrTQhqUntOBvp09qao9G3iw11IeEZGjMNyOzsrMg6KA.A756CCU4_cjoE0qvp "
        self.work_book = ['positionPublishTime', 'positionName', 'pay',	'education', 'positionWorkingExp',
                          'positionWorkCity', 'positionCityDistrict', 'workAddress', 'workType', 'companyName',
                          'financingStageName',	'companySize', 'industryNameLevel',	'welfare_values', 'skill_values',
                          'companyUrl',	'positionUrl', 'companyDescription', 'jobDesc']
        self.count = 1
        self.row_list = []
        self.col_list = []
        self.number = []
        self.item = {}
        self.workbook = Workbook()
        self.work = self.workbook.active
        self.work.append(self.work_book)

    def start_requests(self, page_num):
        for page in range(1, page_num + 1):
            yield self.url.format(page)

    def get_content(self, page_Num):
        for url in self.start_requests(page_num=page_Num):
            try:
                print(url)
                value = random.choice(self.user_agent)
                self.headers['user_agent'] = value
                response = requests.get(url=url, headers=self.headers)
                print(response.status_code)
                if response.status_code == 200:
                    # print(response.text)
                    number = re.findall('"number":"(.*?)"', response.text, re.S)
                    self.number.append(number)
                    print(number)
                else:
                    print('--' * 10 + "当前获取的url状态码为：", response.status_code)

            except:
                print('--'* 10 + '连接url失败！' + '--' * 10)

    def get_json(self):
        global response
        for number in self.number:
            value_1 = random.choice(self.user_agent)
            self.headers['user_agent'] = value_1
            if number:
                for value in number:
                    try:
                        response = requests.get(url=self.json_url.format(value), headers=self.headers)
                        if response.status_code == 200:
                            print('--' * 10 + '正在爬取' + value + '的json数据' + '--' * 10 + str(self.count))
                            self.count += 1
                            json_text = json.loads(response.text)

                            positionPublishTime = json_text['data']['detailedPosition']['positionPublishTime']
                            self.item['positionPublishTime'] = positionPublishTime
                            positionName = json_text['data']['detailedPosition']['positionName']
                            self.item['positionName'] = positionName
                            salary60 = json_text['data']['detailedPosition']['salary60']
                            self.item['pay'] = salary60
                            education = json_text['data']['detailedPosition']['education']
                            self.item['education'] = education
                            positionWorkingExp = json_text['data']['detailedPosition']['positionWorkingExp']
                            self.item['positionWorkingExp'] = positionWorkingExp
                            positionWorkCity = json_text['data']['detailedPosition']['positionWorkCity']
                            self.item['positionWorkCity'] = positionWorkCity
                            positionCityDistrict = json_text['data']['detailedPosition']['positionCityDistrict']
                            self.item['positionCityDistrict'] = positionCityDistrict
                            workAddress = json_text['data']['detailedPosition']['workAddress']
                            self.item['workAddress'] = workAddress
                            workType = json_text['data']['detailedPosition']['workType']
                            self.item['workType'] = workType
                            companyName = json_text['data']['detailedCompany']['companyName']
                            self.item['companyName'] = companyName
                            financingStageName = json_text['data']['detailedCompany']['financingStageName']
                            self.item['financingStageName'] = financingStageName
                            companySize = json_text['data']['detailedCompany']['companySize']
                            self.item['companySize'] = companySize
                            industryNameLevel = json_text['data']['detailedCompany']['industryNameLevel']
                            self.item['industryNameLevel'] = industryNameLevel
                            welfareLabel = json_text['data']['detailedPosition']['welfareLabel']
                            # print(welfareLabel)
                            welfare_values = re.findall("'value': '(.*?)'", str(welfareLabel), re.S)
                            # print(welfare_values)
                            self.item['welfare_values'] = '{}'.format(welfare_values)
                            skillLabel = json_text['data']['detailedPosition']['skillLabel']
                            # print(skillLabel)
                            skill_values = re.findall("'value': '(.*?)'", str(skillLabel), re.S)
                            # print(skill_values)
                            self.item['skill_values'] = '{}'.format(skill_values)
                            companyUrl = json_text['data']['detailedCompany']['companyUrl']
                            self.item['companyUrl'] = companyUrl
                            positionUrl = json_text['data']['detailedPosition']['positionUrl']
                            self.item['positionUrl'] = positionUrl
                            companyDescription = json_text['data']['detailedCompany']['companyDescription']
                            self.item['companyDescription'] = companyDescription
                            jobDesc = json_text['data']['detailedPosition']['jobDesc']
                            pattern = re.compile(r'<[^>]+>', re.S)  # 把不必要的标签字符剔除
                            job_text = pattern.sub('', jobDesc)
                            self.item['jobDesc'] = job_text

                            # companyNumber = json_text['data']['detailedCompany']['companyNumber']
                            # print(companyName, financingStageName, industryNameLevel)
                            # print(companyDescription)
                            # print(workType)
                            # positionHighlight = json_text['data']['detailedPosition']['positionHighlight']
                            # print(workAddress, positionPublishTime)
                            # print(education, positionUrl, salary60, positionWorkingExp)
                            # print(positionWorkCity, positionName, positionCityDistrict)
                            # print(job_text)
                            self.write_excel(item=self.item)
                            self.write_mongodb(item=self.item)
                            # print(self.item)
                        else:
                            print('当前获取的json_url状态码：', response.status_code)
                    except:
                        print('--' * 10 + 'url连接失败:', response.status_code)
            else:
                continue


    def write_excel(self, item):
        # print(item)
        item_values = list(item.values())
        self.work.append(item_values)

    def set_excel(self):
        for v in range(65, 65 + len(self.item)):
            self.row_list.append(chr(v))
        for v_1 in self.row_list:
            self.col_list.append(v_1 + '{}')

        for row_height in range(1,  self.work.max_row + 1):
            self.work.row_dimensions[row_height].height = 40

        for col_width in self.row_list:
            self.work.column_dimensions[col_width].width = 50

        align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for value in self.col_list:
            for num in range(1, self.work.max_row + 1):
                self.work[value.format(num)].alignment = align

    def write_mongodb(self, item):
        client = pymongo.MongoClient('127.0.0.1', 27017)
        db = client['my_mongo']
        db['zp'].insert(dict(item))
        client.close()

    def run(self, page_number):
        self.get_content(page_number)
        self.get_json()
        self.set_excel()
        self.workbook.save('./智联招聘.xlsx')
        self.workbook.close()


if __name__ == '__main__':
    ZL = ZhiLian()
    page_number = int(input('请输入页码:'))
    # keyword = input('请输入关键字:')
    ZL.run(page_number)


